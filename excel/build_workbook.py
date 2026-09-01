import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import os

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
club = pd.read_csv(os.path.join(BASE, "data", "messi_club_stats.csv"))
intl = pd.read_csv(os.path.join(BASE, "data", "messi_international_stats.csv")).fillna("")

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="A50044")
HEADER_FONT = Font(name="Arial", color="FFFFFF", bold=True)
BASE_FONT = Font(name="Arial", size=11)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_table(ws, df, start_row=1):
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=col.replace("_", " ").title())
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BASE_FONT
            c.border = BORDER
    for j, col in enumerate(df.columns, start=1):
        width = max(len(col), df[col].astype(str).map(len).max()) + 3
        ws.column_dimensions[get_column_letter(j)].width = width
    return start_row + len(df) + 1  # next free row

# ---------------- Sheet 1: Club Stats (raw + formula summary) ----------------
ws1 = wb.active
ws1.title = "Club Stats"
next_row = write_table(ws1, club)
n = len(club)
first_data_row = 2
last_data_row = first_data_row + n - 1

goals_col = get_column_letter(list(club.columns).index("goals") + 1)
assists_col = get_column_letter(list(club.columns).index("assists") + 1)
apps_col = get_column_letter(list(club.columns).index("appearances") + 1)
trophy_col = get_column_letter(list(club.columns).index("trophies_won") + 1)

summary_row = last_data_row + 3
ws1.cell(row=summary_row, column=1, value="TOTALS").font = Font(name="Arial", bold=True)
labels = ["Total Goals", "Total Assists", "Total Appearances", "Total Trophies", "Goals per Appearance"]
formulas = [
    f"=SUM({goals_col}{first_data_row}:{goals_col}{last_data_row})",
    f"=SUM({assists_col}{first_data_row}:{assists_col}{last_data_row})",
    f"=SUM({apps_col}{first_data_row}:{apps_col}{last_data_row})",
    f"=SUM({trophy_col}{first_data_row}:{trophy_col}{last_data_row})",
    f"=ROUND(SUM({goals_col}{first_data_row}:{goals_col}{last_data_row})/SUM({apps_col}{first_data_row}:{apps_col}{last_data_row}),2)",
]
for i, (lab, f) in enumerate(zip(labels, formulas)):
    ws1.cell(row=summary_row + 1 + i, column=1, value=lab).font = BASE_FONT
    ws1.cell(row=summary_row + 1 + i, column=2, value=f).font = Font(name="Arial", bold=True)

ws1.freeze_panes = "A2"

# Chart: goals per season
chart = BarChart()
chart.title = "Goals per Club Season"
chart.y_axis.title = "Goals"
chart.x_axis.title = "Season"
data = Reference(ws1, min_col=list(club.columns).index("goals") + 1, min_row=1, max_row=last_data_row)
cats = Reference(ws1, min_col=list(club.columns).index("season") + 1, min_row=2, max_row=last_data_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 24
chart.height = 11
ws1.add_chart(chart, f"A{summary_row + 8}")

# ---------------- Sheet 2: Club Totals by team (SUMIFS-driven) ----------------
ws2 = wb.create_sheet("Club Comparison")
clubs = club["club"].unique().tolist()
headers2 = ["Club", "Seasons", "Total Appearances", "Total Goals", "Total Assists", "Goals per Appearance", "Total Trophies"]
for j, h in enumerate(headers2, start=1):
    c = ws2.cell(row=1, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal="center")

for i, cl in enumerate(clubs, start=2):
    ws2.cell(row=i, column=1, value=cl).font = BASE_FONT
    ws2.cell(row=i, column=2, value=f'=COUNTIF(\'Club Stats\'!{get_column_letter(list(club.columns).index("club")+1)}:{get_column_letter(list(club.columns).index("club")+1)},A{i})').font = BASE_FONT
    ws2.cell(row=i, column=3, value=f'=SUMIF(\'Club Stats\'!{get_column_letter(list(club.columns).index("club")+1)}:{get_column_letter(list(club.columns).index("club")+1)},A{i},\'Club Stats\'!{apps_col}:{apps_col})').font = BASE_FONT
    ws2.cell(row=i, column=4, value=f'=SUMIF(\'Club Stats\'!{get_column_letter(list(club.columns).index("club")+1)}:{get_column_letter(list(club.columns).index("club")+1)},A{i},\'Club Stats\'!{goals_col}:{goals_col})').font = BASE_FONT
    ws2.cell(row=i, column=5, value=f'=SUMIF(\'Club Stats\'!{get_column_letter(list(club.columns).index("club")+1)}:{get_column_letter(list(club.columns).index("club")+1)},A{i},\'Club Stats\'!{assists_col}:{assists_col})').font = BASE_FONT
    ws2.cell(row=i, column=6, value=f"=ROUND(D{i}/C{i},2)").font = BASE_FONT
    ws2.cell(row=i, column=7, value=f'=SUMIF(\'Club Stats\'!{get_column_letter(list(club.columns).index("club")+1)}:{get_column_letter(list(club.columns).index("club")+1)},A{i},\'Club Stats\'!{trophy_col}:{trophy_col})').font = BASE_FONT
    for j in range(1, 8):
        ws2.cell(row=i, column=j).border = BORDER

for j, h in enumerate(headers2, start=1):
    ws2.column_dimensions[get_column_letter(j)].width = max(len(h) + 3, 20)

chart2 = BarChart()
chart2.title = "Total Goals by Club"
chart2.y_axis.title = "Goals"
data2 = Reference(ws2, min_col=4, min_row=1, max_row=len(clubs) + 1)
cats2 = Reference(ws2, min_col=1, min_row=2, max_row=len(clubs) + 1)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width = 16
chart2.height = 10
ws2.add_chart(chart2, "I2")

# ---------------- Sheet 3: International Stats ----------------
ws3 = wb.create_sheet("International Stats")
next_row3 = write_table(ws3, intl)
n3 = len(intl)
last3 = 1 + n3
g_col = get_column_letter(list(intl.columns).index("goals") + 1)
a_col = get_column_letter(list(intl.columns).index("assists") + 1)
tourn_col = get_column_letter(list(intl.columns).index("tournament") + 1)
year_col = get_column_letter(list(intl.columns).index("year") + 1)

# Helper column (non-array): goals only where tournament is FIFA World Cup, else blank
helper_col_idx = len(intl.columns) + 1
helper_col = get_column_letter(helper_col_idx)
ws3.cell(row=1, column=helper_col_idx, value="WC Goals Helper").font = Font(name="Arial", italic=True, size=9, color="999999")
for r in range(2, last3 + 1):
    ws3.cell(row=r, column=helper_col_idx,
             value=f'=IF({tourn_col}{r}="FIFA World Cup",{g_col}{r},"")').font = Font(name="Arial", size=9, color="999999")

sr3 = last3 + 3
ws3.cell(row=sr3, column=1, value="TOTALS (major tournaments only)").font = Font(name="Arial", bold=True)
ws3.cell(row=sr3 + 1, column=1, value="Total Goals").font = BASE_FONT
ws3.cell(row=sr3 + 1, column=2, value=f"=SUM({g_col}2:{g_col}{last3})").font = Font(name="Arial", bold=True)
ws3.cell(row=sr3 + 2, column=1, value="Total Assists").font = BASE_FONT
ws3.cell(row=sr3 + 2, column=2, value=f"=SUM({a_col}2:{a_col}{last3})").font = Font(name="Arial", bold=True)
ws3.cell(row=sr3 + 3, column=1, value="Trophies (Champion results)").font = BASE_FONT
ws3.cell(row=sr3 + 3, column=2, value=f'=COUNTIF(C2:C{last3},"Champion")').font = Font(name="Arial", bold=True)

chart3 = LineChart()
chart3.title = "International Goals per Tournament"
chart3.y_axis.title = "Goals"
data3 = Reference(ws3, min_col=list(intl.columns).index("goals") + 1, min_row=1, max_row=last3)
cats3 = Reference(ws3, min_col=1, min_row=2, max_row=last3)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.width = 22
chart3.height = 10
ws3.add_chart(chart3, f"A{sr3 + 6}")

# ---------------- Sheet 4: Creative Insights (formula-driven, no CSE arrays) ----------------
ws4 = wb.create_sheet("Creative Insights")
ws4.column_dimensions["A"].width = 55
ws4.column_dimensions["B"].width = 16
ws4.column_dimensions["C"].width = 48

title = ws4.cell(row=1, column=1, value="Creative Career Findings")
title.font = Font(name="Arial", bold=True, size=14, color="A50044")

club_col = get_column_letter(list(club.columns).index("club") + 1)
age_col = get_column_letter(list(club.columns).index("approx_age") + 1)
season_col = get_column_letter(list(club.columns).index("season") + 1)
assist_col = get_column_letter(list(club.columns).index("assists") + 1)

rows_def = []

# 1. Age-defying output (SUMIFS — no array needed)
rows_def.append((
    "1. Share of career club goals scored at age 35+",
    f"=ROUND(100*SUMIFS('Club Stats'!{goals_col}:{goals_col},'Club Stats'!{age_col}:{age_col},\">=35\")/SUM('Club Stats'!{goals_col}:{goals_col}),1)&\"%\"",
    "Contradicts the typical athletic decline curve"
))

# 2. Consistency (CV) for Barcelona — built from SUMPRODUCT, no CSE array needed
mean_formula = f"SUMIF('Club Stats'!{club_col}:{club_col},\"FC Barcelona\",'Club Stats'!{goals_col}:{goals_col})/COUNTIF('Club Stats'!{club_col}:{club_col},\"FC Barcelona\")"
rows_def.append((
    "2a. FC Barcelona — average goals per season",
    f"=ROUND({mean_formula},1)",
    ""
))
variance_formula = (
    f"SUMPRODUCT(('Club Stats'!{club_col}2:{club_col}{last_data_row}=\"FC Barcelona\")*"
    f"('Club Stats'!{goals_col}2:{goals_col}{last_data_row}-B4)^2)/"
    f"COUNTIF('Club Stats'!{club_col}:{club_col},\"FC Barcelona\")"
)
rows_def.append((
    "2b. FC Barcelona — goals/season standard deviation",
    f"=ROUND(SQRT({variance_formula}),1)",
    "Reference: B4 = average goals per season (above)"
))
rows_def.append((
    "2c. FC Barcelona — coefficient of variation (stdev / mean)",
    "=ROUND(B5/B4,2)",
    "Lower = more consistent output year to year, despite a 17-season span"
))

# 3. Trophy efficiency by club/country (all SUMIF, no arrays)
rows_def.append(("3a. FC Barcelona — goals per trophy",
                  f"=ROUND(SUMIF('Club Stats'!{club_col}:{club_col},\"FC Barcelona\",'Club Stats'!{goals_col}:{goals_col})/SUMIF('Club Stats'!{club_col}:{club_col},\"FC Barcelona\",'Club Stats'!{trophy_col}:{trophy_col}),1)",
                  "Higher = took more goals to convert into silverware"))
rows_def.append(("3b. Inter Miami CF — goals per trophy",
                  f"=ROUND(SUMIF('Club Stats'!{club_col}:{club_col},\"Inter Miami CF\",'Club Stats'!{goals_col}:{goals_col})/SUMIF('Club Stats'!{club_col}:{club_col},\"Inter Miami CF\",'Club Stats'!{trophy_col}:{trophy_col}),1)",
                  ""))
rows_def.append(("3c. Paris Saint-Germain — goals per trophy",
                  f"=ROUND(SUMIF('Club Stats'!{club_col}:{club_col},\"Paris Saint-Germain\",'Club Stats'!{goals_col}:{goals_col})/SUMIF('Club Stats'!{club_col}:{club_col},\"Paris Saint-Germain\",'Club Stats'!{trophy_col}:{trophy_col}),1)",
                  ""))
rows_def.append(("3d. Argentina (International) — goals per trophy",
                  f"=ROUND(SUM('International Stats'!{g_col}2:{g_col}{last3})/COUNTIF('International Stats'!C2:C{last3},\"Champion\"),1)",
                  "Lowest of all four = his goals converted to trophies MOST efficiently for Argentina"))

# 4. Assist share evolution — SUMPRODUCT-based, no CSE array needed
def season_conditions(seasons):
    return "+".join(f"('Club Stats'!{season_col}2:{season_col}{last_data_row}=\"{s}\")" for s in seasons)

early_conditions = season_conditions(["2004-05", "2005-06", "2006-07", "2007-08"])
late_conditions = season_conditions(["2024", "2025", "2026"])

rows_def.append((
    "4a. Avg. assist share of goal contributions — early career (2004-08)",
    (f"=ROUND(100*SUMPRODUCT(({early_conditions})"
     f"*('Club Stats'!{assist_col}2:{assist_col}{last_data_row}/('Club Stats'!{goals_col}2:{goals_col}{last_data_row}+'Club Stats'!{assist_col}2:{assist_col}{last_data_row})))"
     f"/SUMPRODUCT(({early_conditions})*1),1)&\"%\""),
    "Evolution from primary scorer to creator"
))
rows_def.append((
    "4b. Avg. assist share of goal contributions — final seasons (2024-26)",
    (f"=ROUND(100*SUMPRODUCT(({late_conditions})"
     f"*('Club Stats'!{assist_col}2:{assist_col}{last_data_row}/('Club Stats'!{goals_col}2:{goals_col}{last_data_row}+'Club Stats'!{assist_col}2:{assist_col}{last_data_row})))"
     f"/SUMPRODUCT(({late_conditions})*1),1)&\"%\""),
    ""
))

# 5. World Cup arc — best World Cup by goals, via helper column + INDEX/MATCH (no CSE)
rows_def.append((
    "5. Highest-scoring World Cup (year)",
    f'=INDEX(\'International Stats\'!{year_col}2:{year_col}{last3},MATCH(MAX(\'International Stats\'!{helper_col}2:{helper_col}{last3}),\'International Stats\'!{helper_col}2:{helper_col}{last3},0))',
    "His best World Cup by goals was his LAST one — 2026, age 39"
))

row_i = 3
for label, formula, note in rows_def:
    c1 = ws4.cell(row=row_i, column=1, value=label)
    c1.font = Font(name="Arial", bold=True)
    c2 = ws4.cell(row=row_i, column=2, value=formula)
    c2.font = Font(name="Arial", color="A50044", bold=True)
    c3 = ws4.cell(row=row_i, column=3, value=note)
    c3.font = Font(name="Arial", italic=True, size=9, color="666666")
    row_i += 1

out_path = os.path.join(BASE, "excel", "Messi_Legacy_Analytics.xlsx")
wb.save(out_path)
print("Saved:", out_path)
